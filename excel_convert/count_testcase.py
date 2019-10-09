#!/usr/bin/env python3.7 
# -*- coding:utf-8 -*-
# Author: Lancer  2019-09-20 10:47:58

import  os,sys,re,psutil,copy
from PyQt5 import  QtGui, QtWidgets,QtCore
from PyQt5.QtWidgets import *
from PyQt5.QtChart import *
from PyQt5.QtCore  import *
import  openpyxl  
from    openpyxl.styles import Font
from  report_paraser  import *
import css

linepattern  = re.compile(r'LC+.+\(')
cfilepattern = re.compile(r'\w+.\.c')
functionpattern = re.compile(r'\w+.\s')
notDumppattern  = re.compile(r'\w+.\S')

class  Savedata(object):
    def __init__(self):
        self.index = []
        self.cname = []
        self.functionname = []
        self.cppfilename  = []
        self.LC = []
        self.SC = []
        self.BC = []
        self.FC = []
        self.PC = []
        self.DC = []
        self.SCC =[] 
        self.MCDC = []
        self.testcasenum = []
        self.NA_name = []
        self.NA_num = 0 
        self.total_case_num = 0
    def  clear(self):
        self.index.clear()
        self.cname.clear()
        self.functionname.clear()
        self.cppfilename.clear()
        self.LC.clear()
        self.SC.clear()
        self.BC.clear()
        self.FC.clear()
        self.PC.clear()
        self.DC.clear()
        self.SCC.clear()
        self.MCDC.clear()
        self.testcasenum.clear()
        self.NA_name.clear()
        self.NA_num = 0
        self.total_case_num = 0

    def  mprint(self):
        for i in range(len(self.cname)):
            print("cname: %s  函数: %s  LC:%s SC:%s BC:%s  FC:%s  PC:%s  DC:%s SCC:%s MCDC:%s  案例个数:%d  "%(self.cname[i],self.functionname[i],self.LC[i], self.SC[i],
            self.BC[i], self.FC[i], self.PC[i], self.DC[i], self.SCC[i], self.MCDC[i],self.testcasenum[i]  ))
            self.total_case_num = self.total_case_num +  self.testcasenum[i]
            if "N/A" == self.MCDC[i] :
                self.NA_num = self.NA_num + 1 
                self.NA_name.append(self.functionname[i])
        if 0 != self.NA_num :
            print("N/A 函数个数:%d 总测试案例个数:%d  如下:"%(self.NA_num,self.total_case_num  ))
            for i in range(self.NA_num):
                print(self.NA_name[i])
        else:
            print("总测试案例个数:%d ",self.total_case_num)



class CountTestCase(QWidget):
    def  __init__(self, parent=None):
        super(CountTestCase,self).__init__(parent)
        self.mainlayout =  QVBoxLayout()
        self.frame_btn =  QFrame()
        self.frame_barstack =  QFrame()
        self.btn_layout = QHBoxLayout()
        self.report_choice_btn = QPushButton("测试报告文档选择(.pdf)")
        self.case_show_btn = QPushButton("查询具体函数案例个数")
        self.excel_create = QPushButton("详细设计单元测试追溯表生成")
        self.btn_layout.addWidget(self.report_choice_btn)
        self.btn_layout.addWidget(self.case_show_btn)
        self.btn_layout.addWidget(self.excel_create)
        self.frame_btn.setLayout(self.btn_layout)
        self.mainlayout.addWidget(self.frame_barstack)
        self.mainlayout.addWidget(self.frame_btn)
        self.mainlayout.setStretchFactor(self.frame_barstack, 9)
        self.mainlayout.setStretchFactor(self.frame_btn, 1)
        #  文件列表显示
        self.filelist_layout = QVBoxLayout()
        # self.left_Widget = QTableWidget()
        self.right_listWidget = QListWidget()
        self.filedirbtn_layout = QHBoxLayout()
        self.list_widget_layout = QHBoxLayout()
        #self.old_label = QLabel("函数案例统计情况")
        self.new_label = QLabel("函数测试文件不存在")
        #self.old_label.setAlignment( QtCore.Qt.AlignCenter)
        self.new_label.setAlignment( QtCore.Qt.AlignCenter)
        #self.filedirbtn_layout.addWidget(self.old_label)
        self.filedirbtn_layout.addWidget(self.new_label)
        self.filedirframe = QFrame()
        # self.list_widget_layout.addWidget(self.left_Widget)
        self.list_widget_layout.addWidget(self.right_listWidget)
        #self.left_listWidget.addItems()
        self.filelist_layout.addSpacing(0)
        self.filelist_layout.addLayout(self.filedirbtn_layout)
        self.filelist_layout.addLayout(self.list_widget_layout )
        self.frame_barstack.setLayout(self.filelist_layout)
        self.filedirbtn_layout.setContentsMargins(0,0,0,0)


        #self.mainlayout.addStretch(1)
        self.setLayout(self.mainlayout)
        #信号槽连接
        self.excel_create.clicked.connect(self.setChart_view)
        self.report_choice_btn.clicked.connect(self.report_choice)
        self.case_show_btn.clicked.connect(self.show_case_data)
        # 数据存储
        self.fileName_choose = ""
        self.fault_cpp_list = []
        self.pubdb = Savedata()
        self.setStyleSheet(css.black)
    
    def  parser_percent_data(self, src, recv_dst ):
        rpers = src.split(r' ') #空格
        VLC   = rpers[0].split("=")
        VMCDC = rpers[7].split("=")
        if("LC" == VLC[0]) and (VMCDC[0] == "MCDC"):
            VSC = rpers[1].split("=")
            VBC = rpers[2].split("=")
            VFC = rpers[3].split("=")
            VPC = rpers[4].split("=")
            VDC = rpers[5].split("=")
            VSCC = rpers[6].split("=")
            #fill data
            recv_dst.LC.append(str(VLC[1]))
            recv_dst.SC.append(str(VSC[1]))
            recv_dst.BC.append(str(VBC[1]))
            recv_dst.FC.append(str(VFC[1]))
            recv_dst.PC.append(str(VPC[1]))
            recv_dst.DC.append(str(VDC[1]))
            recv_dst.SCC.append(str(VSCC[1]))
            recv_dst.MCDC.append(str(VMCDC[1]))
        else:
            print("LC ··· MCDC format  err")
    # 测试案例编号
    def  count_test_case_num(self,functionname, cppfilename):
        case_num = 0 
        functionname = functionname.strip()
        with open(cppfilename,encoding='gbk') as f:
            lines = f.readlines()
            for  line in lines:
                find = re.findall(r"测试案例编号",line)
                if len(find) > 0 :
                    case_num = case_num + len(find)
        return  (case_num)

    def  create_excel(self,pubdb):
        wb = openpyxl.Workbook()
        sheet = wb.get_active_sheet()
        #wb.remove_sheet("sheet")
        #sheet = wb.create_sheet("模块详细设计与单元测试函数的追溯表")
        sheet['A1'] = '模块'
        sheet['B1'] = '详细设计文档编号'
        sheet['C1'] = '文件'
        sheet['D1'] = '函数'
        sheet['E1'] = '状态'
        sheet['F1'] = '用例数'
        sheet['G1'] = '断言失败'
        sheet['H1'] = '测试用例代码'
        sheet['I1'] = '行覆盖率'
        sheet['J1'] = '语句覆盖率'
        sheet['K1'] = '基本块覆盖率'
        sheet['L1'] = '函数覆盖率'
        sheet['M1'] = '路径覆盖率'
        sheet['N1'] = '判定覆盖率'
        sheet['O1'] = '简单条件覆盖率'
        sheet['P1'] = '修改条件/判定覆盖率'
        fontObj = Font(name='Segoe UI',size=9,bold=True)

        for  i  in  range(len(pubdb.cname)):
            #行 列
            sheet.cell(row = i+2, column = 3).value =  pubdb.cname[i]  #文件
            sheet.cell(row = i+2, column = 4).value =  pubdb.functionname[i]  #函数
            sheet.cell(row = i+2, column = 5).value =  "OK"  #状态
            sheet.cell(row = i+2, column = 6).value =  pubdb.testcasenum[i]  #用例数
            sheet.cell(row = i+2, column = 7).value =  "0"  #断言失败
            sheet.cell(row = i+2, column = 8).value =  pubdb.cppfilename[i]  #测试用例代码
            sheet.cell(row = i+2, column = 9).value =  pubdb.LC[i]  #行覆盖率
            sheet.cell(row = i+2, column = 10).value =  pubdb.SC[i]  #语句覆盖率
            sheet.cell(row = i+2, column = 11).value = pubdb.BC[i]  #基本块覆盖率
            sheet.cell(row = i+2, column = 12).value = pubdb.FC[i]  #函数覆盖率
            sheet.cell(row = i+2, column = 13).value = pubdb.PC[i]  #路径覆盖率
            sheet.cell(row = i+2, column = 14).value = pubdb.DC[i]  #判定覆盖率
            sheet.cell(row = i+2, column = 15).value = pubdb.SCC[i]  #简单条件覆盖率
            sheet.cell(row = i+2, column = 16).value = pubdb.MCDC[i]  #修改条件/判定覆盖率
        wb.save('LRTSW-CI-子系统模块详细设计与单元测试函数的追溯表.xlsx')

    #测试报告文件选择
    def  report_choice(self):
        self.fileName_choose = ""
        self.fileName_choose, filetype = QFileDialog.getOpenFileName(self,  
                                    "选取文件",  
                                    os.getcwd(), # 起始路径 
                                    "Text Files (*.pdf)")   # 设置文件扩展名过滤,用双分号间隔

        print("\n你选择的文件为:")
        if self.fileName_choose == "":
            print("错误:请未选择报告")
            return 
        print(self.fileName_choose)
        self.fault_cpp_list.clear()
        # 读取报告内容
        self.pubdb.clear()
        #with open(self.fileName_choose,encoding='gbk') as f:
        #paraser_report_pdf
        lines = paraser_report_pdf(self.fileName_choose)  
        lines = get_rate_percent(lines)
        #print(lines)
        file = open("report_01.txt", 'w',encoding='utf-8')
        file.write(lines)
        file.close()
        
        if True == os.path.exists('report_01.txt'):
            with open('report_01.txt',encoding='utf-8') as f:
                lines = f.readlines() 
        os.remove("report_01.txt")
        lines = str(lines)
        lines = lines.replace(r"\xa0", " ")
        lines = lines.replace(r"\n", " ")
        lines = lines.split(",")
        tfilename = ""
        for i  in  range(len(lines)):
            if 0 == len(lines[i]):
                lines.remove(lines[i])
            else:
                lines[i] =  lines[i].strip() +"]"
        for i  in  range(len(lines)):
            #匹配.c文件
            line = lines[i].strip() 
            line = line.replace(r"\xa0", " ")
            cfilename = cfilepattern.search(line)
            if cfilename is not None:
                tfilename = cfilename.group()
            else:
                self.pubdb.cname.append(tfilename)  #.c  name
                #print()
                functionname =  functionpattern.search(line)  #function name
                self.pubdb.functionname.append(functionname.group())  #.c  name
                functionname =  notDumppattern.search(functionname.group()) #notDumppattern
                functionname = functionname.group(0)
    
                cppfilename =  "test_"+functionname+".cpp"
                self.pubdb.cppfilename.append(cppfilename)
                #打开测试文件计算 测试案例个数
                if True == os.path.exists(cppfilename):
                    case_num = self.count_test_case_num(functionname, cppfilename )
                    #print("%s  num:%d"%(functionname,case_num ))
                    self.pubdb.testcasenum.append(case_num)
                else:
                    self.fault_cpp_list.append("test_"+functionname+".cpp")
                    #print("test_"+functionname+".cpp"+"  is not exist===>>check")
                    self.pubdb.testcasenum.append(0)  #test case num
                
                #cppfilename
                notcfileline = linepattern.search(line)
                percentline = notcfileline.group()
                percentline  = percentline.split('(')
                #print(percentline)
                self.parser_percent_data( percentline[0], self.pubdb)
        self.pubdb.mprint()
        

    def  show_case_data(self):
        show  = QDialog()
        show.setModal(True)
        show.setWindowTitle("案例统计情况")
        hhh = QHBoxLayout()
        left_Widget = QTableWidget()
        hhh.addWidget(left_Widget)
        show.setLayout(hhh)
        #N列 2行
        left_Widget.setColumnCount(2)
        left_Widget.setRowCount(len(self.pubdb.cname))
        left_Widget.setHorizontalHeaderLabels(["函数名", "案例个数"])
        left_Widget.setColumnWidth(0,400) #设置0列的宽度
        left_Widget.setColumnWidth(1,100)
        show.setFixedWidth(600)
        show.setStyleSheet(css.black)
        
        for i in range(len(self.pubdb.cname)):
            left_Widget.setRowHeight(i,40)
            left_Widget.setItem(i, 0, QTableWidgetItem(self.pubdb.functionname[i]))    #设置j行i列的内容为Value
            left_Widget.setItem(i, 1, QTableWidgetItem(str(self.pubdb.testcasenum[i] ))) #设置j行i列的内容为Value
        if len(self.pubdb.cname)*40+80 <= 800:
            show.setFixedHeight(len(self.pubdb.cname)*40+80)
        else:
            show.setFixedHeight(800)
        show.show()
        show.exec_()
        pass
        
    #显示报告中有测试案例统计失败函数
    def  setChart_view(self):
        self.right_listWidget.clear()
        self.right_listWidget.addItems(self.fault_cpp_list)
        self.create_excel(self.pubdb)
        #self.fault_cpp_list.clear()


        
    
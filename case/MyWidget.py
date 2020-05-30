#!/usr/bin/env python
# encoding: utf-8
# @Time    : 2020/4/14 10:46
# @Author  : Lancer

import sys,os,re,docx,threading,time
import  openpyxl
from PyQt5 import QtCore, QtGui
from PyQt5.QtWidgets import *
from show_style import *
from  report_paraser import *
#from  myThread  import *


__VERSION__ = "版本号:1.0.0"
__AUTHOR__ = " 作者:Lancer "
__TIME_M__ = " 2020-04-04"

label_text = \
"""
 使用说明:

1.点击按钮 单元测试报告  选择单元测试报告.pdf
2.点击按钮 详细设计报告  选择 详细设计报告.docx
3.点击按钮 显示追溯信息  查看单元测试 详细设计
  函数追溯信息
4.点击按钮  创建追溯表         
  生成追溯表格
"""

pdfread_flag = 0
read_finish  = 0 
pdf_path = ""
threadLock = threading.Lock()

class myThread (threading.Thread):
    def __init__(self):
        threading.Thread.__init__(self)

    def run(self):
        
        global pdfread_flag
        global read_finish
        while True:
            try:
                if pdfread_flag == 1 :
                    starttime = time.time()
                    print("开始时间:", starttime )
                    with pdfplumber.open(pdf_path) as pdf:
                        pdf_page_num = len(pdf.pages)
                        print("报告页数: %d" % pdf_page_num)
                        text = ""
                        for i in range(pdf_page_num):
                            text_temp = pdf.pages[i].extract_text()
                            if text_temp:
                                text = text + text_temp
                        pdf.close()
                    endtime = time.time()
                    print("结束时间:", endtime )
                    print("读取解析PDF时间",endtime-starttime )
                    threadLock.acquire()
                    pdfread_flag = 0 
                    read_finish = 1
                    threadLock.release()
                else:
                    read_finish = 0
            except Exception as e:
                pass
            time.sleep(1)


class MyWidget(QWidget):
    def __init__(self):
        super(MyWidget,self).__init__()
        self.resize(1000,600)
        self.setWindowTitle("Tools"+"   "+ __VERSION__ + " " + __AUTHOR__ + __TIME_M__ )
        mailayout = QHBoxLayout()
        btnlayout = QVBoxLayout()
        display_layout = QVBoxLayout()
        mailayout.addLayout(display_layout,3)
        mailayout.addLayout(btnlayout, 1)
        self.pdf_chose_btn =  QPushButton("单元测试报告")
        self.md_chose_btn = QPushButton("详细设计报告")
        self.create_excel_btn = QPushButton("创建追溯表")
        self.check_btn = QPushButton("显示追溯信息")
        self.help_btn = QPushButton("帮助信息")
        btnlayout.addWidget(self.pdf_chose_btn)
        btnlayout.addWidget(self.md_chose_btn)
        btnlayout.addWidget(self.create_excel_btn)
        btnlayout.addWidget(self.check_btn)
        btnlayout.addWidget(self.help_btn)
        btnlayout.addStretch()
        self.text_show = QTextEdit()
        display_layout.addWidget(self.text_show)
        self.setLayout(mailayout)
        self.setStyleSheet(style_format_01)

        self.pdf_chose_btn.clicked.connect(self.slot_pdf_chose_btn)
        self.md_chose_btn.clicked.connect(self.slot_md_chose_btn)
        self.create_excel_btn.clicked.connect(self.slot_create_excel_btn)
        self.check_btn.clicked.connect(self.slot_check_btn)
        self.help_btn.clicked.connect(self.slot_help_btn)

        self.pdf_obj = ParaserPdf()
        self.detail_designer_fun = []

        self.child_threading_pdf = myThread()
        self.child_threading_pdf.setDaemon(True)   #把子进程设置为守护线程，必须在start()之前设置
        
        self.child_threading_pdf.start()
        ###定时器
        self.timer= QtCore.QTimer()
        self.timer.start(1000)
        self.timer.timeout.connect(self.slot_cyc_up_data_ui)
        self.last_v = 0

    def slot_cyc_up_data_ui(self):
        global read_finish
        if read_finish == 1 :
            if  0 != len(self.pdf_obj.cpp_notfind_list):
                self.text_show.setText("PDF 测试报告读取解析成功\n 未找到测试文件函数:\n"+"\n".join(self.pdf_obj.cpp_notfind_list))
            else:
                self.text_show.setText("PDF 测试报告读取解析成功  所有函数都找到对应测试文件\n ")



    def slot_help_btn(self):
        QMessageBox.about(self, "帮助",label_text)


    def slot_pdf_chose_btn(self):
        self.fileName_choose = ""
        global pdf_path
        self.text_show.setText("请等待 PDF 测试报告读取解析中···")
        self.fileName_choose, filetype = QFileDialog.getOpenFileName(self,
                                                                     "选取文件",
                                                                     os.getcwd(),  # 起始路径
                                                                     "Text Files (*.pdf)")  # 设置文件扩展名过滤,用双分号间隔
        if self.fileName_choose == "":
            self.text_show.setText("错误:未选择报告  请重新点击按钮 选择单元测试报告")
            return
        pdf_path = self.fileName_choose
        global pdfread_flag
        threadLock.acquire()
        pdfread_flag = 1
        threadLock.release()
        # test = self.pdf_obj.paraser_report_pdf(self.fileName_choose)
        # self.pdf_obj.paraser_and_fill_dt(test)
        #self.text_show.setText("PDF 测试报告读取解析成功\n 未找到测试文件函数:\n"+"\n".join(self.pdf_obj.cpp_notfind_list))


    def slot_md_chose_btn(self):
        fileName_choose = ""
        self.text_show.setText("详细设计 word 读取解析中···")
        fileName_choose, filetype = QFileDialog.getOpenFileName(self,
                                                                     "选取文件",
                                                                     os.getcwd(),  # 起始路径
                                                                     "Text Files (*.docx)")  # 设置文件扩展名过滤,用双分号间隔
        if  not fileName_choose:
            self.text_show.setText("错误:未选择报告  请重新点击按钮 选择详细设计报告")
            return
        
        self.detail_designer_fun.clear()
        self.__get_function_name(fileName_choose)
        self.text_show.setText("详细设计 word 读取解析成功")

    def __get_function_name(self,mdfile:str):
        try:
            doc = docx.Document(mdfile)
            content = '\n'.join([para.text for para in doc.paragraphs])
            fun_pattern = re.compile('函数.*流程图')
            mobj = fun_pattern.findall(content)
            for c in mobj:
                c = c.replace("函数","").replace("流程图","")
                self.detail_designer_fun.append(c)
        finally:
            pass
            #doc.close()
        # try:
        #     file = open(mdfile,mode="r+",encoding="utf-8")
        # except  Exception as e:
        #     file = open(mdfile, mode="r+", encoding="gbk")
        # finally:
        #     lines = file.readlines()
        #     for line in lines:
        #         line_src = line
        #         # print(line)
        #         line = line.strip()
        #         if line.startswith("## 子函数"):
        #             self.detail_designer_fun.append(line.split("## 子函数")[-1])
        #     file.close()


    def slot_check_btn(self):
        self.text_show.setText("详细设计  单元测试报告函数一致")
        str_ret = ""
        function_list = []
        for i in range(len(self.pdf_obj.excel_data)):
            function_list.append(self.pdf_obj.excel_data[i][1])

        for name in  function_list:
            if name not in self.detail_designer_fun:
                str_ret = str_ret + name + "\n"
        if str_ret:
            str_ret = "\n单元测试报告出现  详细设计未出现函数:\n" + str_ret 
            str_ret = str_ret + "\n详细设计出现 单元测试未出现函数\n"


        for name in  self.detail_designer_fun:
            if name not in  function_list:
                str_ret = str_ret + name + "\n"
        if  str_ret !="":
            self.text_show.setText(str_ret)
        else:
            self.text_show.setText("单元测试函数 详细设计函数校验一致")

    def  slot_create_excel_btn(self):
        wb = openpyxl.Workbook()
        sheet = wb.active
        # wb.remove_sheet("sheet")
        sheet['A1'] = '模块'
        sheet['B1'] = '详细设计文档编号'
        sheet['C1'] = '文件'
        sheet['D1'] = '函数'
        sheet['E1'] = '状态'
        sheet['F1'] = '用例数'
        sheet['G1'] = '断言失败'
        sheet['H1'] = '测试用例代码'
        sheet['I1'] = '行覆盖率'
        sheet['J1'] = '语句覆盖率'
        sheet['K1'] = '基本块覆盖率'
        sheet['L1'] = '函数覆盖率'
        sheet['M1'] = '路径覆盖率'
        sheet['N1'] = '判定覆盖率'
        sheet['O1'] = '简单条件覆盖率'
        sheet['P1'] = '修改条件/判定覆盖率'


        #self.notfind_label.setText("未找到函数测试文件个数: %d" % self.fault_cpp_num)
        total_case_num = 0
        print(len(self.pdf_obj.excel_data))
        for i in range(len(self.pdf_obj.excel_data)):
            # 行 列
            sheet.cell(row=i + 2, column=3).value = self.pdf_obj.excel_data[i][0]  # 文件
            sheet.cell(row=i + 2, column=4).value = self.pdf_obj.excel_data[i][1]  # 函数
            sheet.cell(row=i + 2, column=5).value = "OK"  # 状态
            sheet.cell(row=i + 2, column=6).value = self.pdf_obj.excel_data[i][10]  # 用例数
            sheet.cell(row=i + 2, column=7).value = "0"  # 断言失败
            sheet.cell(row=i + 2, column=8).value = "test_" + self.pdf_obj.excel_data[i][1] + ".cpp"  # 测试用例代码
            sheet.cell(row=i + 2, column=9).value = self.pdf_obj.excel_data[i][2]  # 行覆盖率
            sheet.cell(row=i + 2, column=10).value = self.pdf_obj.excel_data[i][3]  # 语句覆盖率
            sheet.cell(row=i + 2, column=11).value = self.pdf_obj.excel_data[i][4]  # 基本块覆盖率
            sheet.cell(row=i + 2, column=12).value = self.pdf_obj.excel_data[i][5]  # 函数覆盖率
            sheet.cell(row=i + 2, column=13).value = self.pdf_obj.excel_data[i][6]  # 路径覆盖率
            sheet.cell(row=i + 2, column=14).value = self.pdf_obj.excel_data[i][7]  # 判定覆盖率
            sheet.cell(row=i + 2, column=15).value = self.pdf_obj.excel_data[i][8]  # 简单条件覆盖率
            sheet.cell(row=i + 2, column=16).value = self.pdf_obj.excel_data[i][9]  # 修改条件/判定覆盖率
            if (-1 != self.pdf_obj.excel_data[i][10]):
                total_case_num = total_case_num + int(self.pdf_obj.excel_data[i][10])
        wb.save('LRTSW-CI-子系统模块详细设计与单元测试函数的追溯表.xlsx')
        wb.close()
        #self.case_num_label.setText("测案例总个数统计: %d" % total_case_num)

if __name__ == "__main__":
    import sys
    app  = QApplication(sys.argv)
    demo = MyWidget()
    demo.show()
    sys.exit(app.exec_())

# -*- coding: utf-8 -*-
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import *
from  tools  import *
import sys,os,re
import  templatestr , getcfunction
from  report_paraser  import *
import  openpyxl
from    openpyxl.styles import Font


class MyTools(Ui_tools):
    def __init__(self):
        super(MyTools,self).__init__()
        self.cfunction_num = 0
        self.cfile_choose = ""
        self.function_name_list = []
        self.fileName_choose = ""
        self.fault_cpp_list = []
        self.fault_cpp_num = 0
        #slot
        self.excel_create_btn.clicked.connect(self.create_excel_report_slot)
        self.pdfchose_btn.clicked.connect(self.chose_report_slot)
        self.sigfile_btn.clicked.connect(self.create_sigtestcase_slot)
        self.allfile_btn.clicked.connect(self.create_cfiletestcase_slot)
        self.cfilechose_btn.clicked.connect(self.chose_cfile_slot)
        

    # 实现逻辑
    def chose_cfile_slot(self):
        """选择待处理的.c文件"""
        self.cfile_choose = ""
        Widget = QWidget()
        self.cfile_choose, filetype = QFileDialog.getOpenFileName(Widget,
                                                                  "选取文件",
                                                                  os.getcwd(),
                                                                  "Text Files (*.c)")
        print("\n你选择的文件为:")
        if self.cfile_choose == "":
            print("错误:请未选择.c源文件")
            return
        print(self.cfile_choose)
        # 提取.c中函数原型
        self.function_name_list = getcfunction.get_cfile_function_name(self.cfile_choose)
        self.cfunction_num = len(self.function_name_list)
        self.function_num_label.setText("函数个数: %d" % self.cfunction_num)

    def create_cfiletestcase_slot(self):
        """生成单.c中所有函数测试文件"""
        for i in range(len(self.function_name_list)):
            function_name =  self.function_name_list[i].strip().split("(")[0]
            function_name = function_name.strip().split(" ")[-1].strip()
            print(function_name)
            test_casenum = int(self.n_casenum_spin.value())
            index = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
            case = ""
            descript01 = ""
            descript02 = ""
            for i in range(test_casenum):
                c_name = {"name": function_name, "casenum": index[i]}
                temp_case = (templatestr.current_case % c_name)
                temp_descript01 = (templatestr.descript_case_01 % c_name)
                temp_descript02 = (templatestr.descript_case_02 % c_name)
                case = case + temp_case
                descript01 = descript01 + temp_descript01
                descript02 = descript02 + temp_descript02
            # print(case)
            file = "test_" + function_name + ".cpp"
            name = {"name": function_name, "casenum": index[i], "case": case, "descript01": descript01,
                    "descript02": descript02}
            if False == os.path.exists(file):
                testfile = open(file, "a")
                testfile.write(templatestr.test_template % name)
                testfile.close()
            else:
                pass

    def create_sigtestcase_slot(self):
        """生成单个函数测试文件"""
        function_name = self.funname_lineEdit.text()
        test_casenum = int(self.one_casenum_spin.value())
        index = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
        case = ""
        descript01 = ""
        descript02 = ""
        for i in range(test_casenum):
            c_name = {"name": function_name, "casenum": index[i]}
            temp_case = (templatestr.current_case % c_name)
            temp_descript01 = (templatestr.descript_case_01 % c_name)
            temp_descript02 = (templatestr.descript_case_02 % c_name)
            case = case + temp_case
            descript01 = descript01 + temp_descript01
            descript02 = descript02 + temp_descript02
        # print(case)

        file = "test_" + function_name + ".cpp"
        name = {"name": function_name, "casenum": index[i], "case": case, "descript01": descript01,
                "descript02": descript02}
        if False == os.path.exists(file):
            testfile = open(file, "a")
            testfile.write(templatestr.test_template % name)
            testfile.close()
        else:
            wdiget = QWidget()
            wdiget.setStyleSheet(css)
            QMessageBox.information(wdiget, "新建测试文件", " %s 已经存在" % file)
            print("文件:%s 已经存在" % file)

    def chose_report_slot(self):
        self.fault_cpp_num = 0
        wdiget = QWidget()
        self.fileName_choose = ""
        self.fileName_choose, filetype = QFileDialog.getOpenFileName(wdiget,
                                                                     "选取文件",
                                                                     os.getcwd(),  # 起始路径
                                                                     "Text Files (*.pdf)")  # 设置文件扩展名过滤,用双分号间隔

        print("\n你选择的文件为:")
        if self.fileName_choose == "":
            print("错误:请未选择报告")
            return
        print(self.fileName_choose)
        self.fault_cpp_list.clear()
        # 读取报告内容
        lines = paraser_report_pdf(self.fileName_choose)
        lines = get_rate_percent(lines)
        lines = lines.split("\n")
        paraser_and_fill_dt(lines)


    def create_excel_report_slot(self):
        """生成追溯表"""
        wb = openpyxl.Workbook()
        sheet = wb.get_active_sheet()
        # wb.remove_sheet("sheet")
        # sheet = wb.create_sheet("模块详细设计与单元测试函数的追溯表")
        sheet['A1'] = '模块'
        sheet['B1'] = '详细设计文档编号'
        sheet['C1'] = '文件'
        sheet['D1'] = '函数'
        sheet['E1'] = '状态'
        sheet['F1'] = '用例数'
        sheet['G1'] = '断言失败'
        sheet['H1'] = '测试用例代码'
        sheet['I1'] = '行覆盖率'
        sheet['J1'] = '语句覆盖率'
        sheet['K1'] = '基本块覆盖率'
        sheet['L1'] = '函数覆盖率'
        sheet['M1'] = '路径覆盖率'
        sheet['N1'] = '判定覆盖率'
        sheet['O1'] = '简单条件覆盖率'
        sheet['P1'] = '修改条件/判定覆盖率'
        fontObj = Font(name='Segoe UI', size=9, bold=True)
        
        self.notfind_label.setText("未找到函数测试文件个数: %d"%self.fault_cpp_num)
        total_case_num = 0 
        for i in range(len(excel_data)):
            # 行 列
            sheet.cell(row=i + 2, column=3).value = excel_data[i][0]  # 文件
            sheet.cell(row=i + 2, column=4).value = excel_data[i][1]  # 函数
            sheet.cell(row=i + 2, column=5).value = "OK"  # 状态
            sheet.cell(row=i + 2, column=6).value = excel_data[i][10]  # 用例数
            sheet.cell(row=i + 2, column=7).value = "0"  # 断言失败
            sheet.cell(row=i + 2, column=8).value = "test_" + excel_data[i][1] + ".cpp"  # 测试用例代码
            sheet.cell(row=i + 2, column=9).value = excel_data[i][2]  # 行覆盖率
            sheet.cell(row=i + 2, column=10).value = excel_data[i][3]  # 语句覆盖率
            sheet.cell(row=i + 2, column=11).value = excel_data[i][4]  # 基本块覆盖率
            sheet.cell(row=i + 2, column=12).value = excel_data[i][5]  # 函数覆盖率
            sheet.cell(row=i + 2, column=13).value = excel_data[i][6]  # 路径覆盖率
            sheet.cell(row=i + 2, column=14).value = excel_data[i][7]  # 判定覆盖率
            sheet.cell(row=i + 2, column=15).value = excel_data[i][8]  # 简单条件覆盖率
            sheet.cell(row=i + 2, column=16).value = excel_data[i][9]  # 修改条件/判定覆盖率
            if (-1 != excel_data[i][10]):
                total_case_num = total_case_num + int(excel_data[i][10])
        wb.save('LRTSW-CI-子系统模块详细设计与单元测试函数的追溯表.xlsx')
        wb.close()
        self.case_num_label.setText("测案例总个数统计: %d"%total_case_num)
        #


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    demo = MyTools()
    demo.show()
    sys.exit(app.exec_())
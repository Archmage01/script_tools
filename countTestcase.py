#!/usr/bin/env python3 
# -*- coding:gbk -*-
# Author: yang.gan  2019-3-3 18:52:50

import  sys,os,re, getopt
import  openpyxl  
from    openpyxl.styles import Font
from    openpyxl.styles import colors

'''
    for   count  testcase  num
'''

class  Savedata(object):
    index = []
    cname = []
    functionname = []
    cppfilename  = []
    LC = []
    SC = []
    BC = []
    FC = []
    PC = []
    DC = []
    SCC =[] 
    MCDC = []
    testcasenum = []
    NA_name = []
    NA_num = 0 
    total_case_num = 0

    def __init__(self):
        pass

    def  mprint(self):
        for i in range(len(self.cname)):
            print("cname: %s  函数: %s  LC:%s SC:%s BC:%s  FC:%s  PC:%s  DC:%s SCC:%s MCDC:%s  案例个数:%d  "%(self.cname[i],self.functionname[i],self.LC[i], self.SC[i],
            self.BC[i], self.FC[i], self.PC[i], self.DC[i], self.SCC[i], self.MCDC[i],self.testcasenum[i]  ))
            self.total_case_num = self.total_case_num +  self.testcasenum[i]
            if "N/A" == self.MCDC[i] :
                self.NA_num = self.NA_num + 1 
                self.NA_name.append(self.functionname[i])
        if 0 != self.NA_num :
            print("N/A 函数个数:%d 总测试案例个数:%d  如下:"%(self.NA_num,self.total_case_num  ))
            for i in range(self.NA_num):
                print(self.NA_name[i])
        else:
            print("总测试案例个数:%d ",self.total_case_num)
            pass


linepattern  = re.compile(r'LC+.+\(')
cfilepattern = re.compile(r'\w+.\.c')
functionpattern = re.compile(r'\w+.\s')
notDumppattern  = re.compile(r'\w+.\S')


def  parser_percent_data(src, recv_dst ):
    arrlen = len(recv_dst.cname)
    rpers = src.split(r' ') #空格
    VLC   = rpers[0].split("=")
    VMCDC = rpers[7].split("=")
    if("LC" == VLC[0]) and (VMCDC[0] == "MCDC"):
        VSC = rpers[1].split("=")
        VBC = rpers[2].split("=")
        VFC = rpers[3].split("=")
        VPC = rpers[4].split("=")
        VDC = rpers[5].split("=")
        VSCC = rpers[6].split("=")
        #fill data
        recv_dst.LC.append(str(VLC[1]))
        recv_dst.SC.append(str(VSC[1]))
        recv_dst.BC.append(str(VBC[1]))
        recv_dst.FC.append(str(VFC[1]))
        recv_dst.PC.append(str(VPC[1]))
        recv_dst.DC.append(str(VDC[1]))
        recv_dst.SCC.append(str(VSCC[1]))
        recv_dst.MCDC.append(str(VMCDC[1]))
    else:
        print("LC ··· MCDC format  err")

#count  test case num 待细化处理
def  count_test_case_num(functionname, cppfilename):
    case_num = 0 
    with open(cppfilename,encoding='gbk') as f:
        lines = f.readlines()
        for  line  in  lines:
            find = re.findall(r"测试案例编号",line)
            if len(find) > 0:
                case_num = case_num + len(find)
        f.close()
    return  (case_num)

#创建Excel
def  create_excel(pubdb):
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    #wb.remove_sheet("sheet")
    #sheet = wb.create_sheet("模块详细设计与单元测试函数的追溯表")
    sheet['A1'] = '模块'
    sheet['B1'] = '详细设计文档编号'
    sheet['C1'] = '文件'
    sheet['D1'] = '函数'
    sheet['E1'] = '状态'
    sheet['F1'] = '用例数'
    sheet['G1'] = '断言失败'
    sheet['H1'] = '测试用例代码'
    sheet['I1'] = '行覆盖率'
    sheet['J1'] = '语句覆盖率'
    sheet['K1'] = '基本块覆盖率'
    sheet['L1'] = '函数覆盖率'
    sheet['M1'] = '路径覆盖率'
    sheet['N1'] = '判定覆盖率'
    sheet['O1'] = '简单条件覆盖率'
    sheet['P1'] = '修改条件/判定覆盖率'
    #垂直水平居中

    #给表头设置字体
    fontObj = Font(name='Segoe UI',size=9,bold=True)


    for  i  in  range(len(pubdb.cname)):
        #行 列
        sheet.cell(row = i+2, column = 3).value =  pubdb.cname[i]  #文件
        sheet.cell(row = i+2, column = 4).value =  pubdb.functionname[i]  #函数
        sheet.cell(row = i+2, column = 5).value =  "OK"  #状态
        sheet.cell(row = i+2, column = 6).value =  pubdb.testcasenum[i]  #用例数
        sheet.cell(row = i+2, column = 7).value =  "0"  #断言失败
        sheet.cell(row = i+2, column = 8).value =  pubdb.cppfilename[i]  #测试用例代码
        sheet.cell(row = i+2, column = 9).value =  pubdb.LC[i]  #行覆盖率
        sheet.cell(row = i+2, column = 10).value =  pubdb.SC[i]  #语句覆盖率
        sheet.cell(row = i+2, column = 11).value = pubdb.BC[i]  #基本块覆盖率
        sheet.cell(row = i+2, column = 12).value = pubdb.FC[i]  #函数覆盖率
        sheet.cell(row = i+2, column = 13).value = pubdb.PC[i]  #路径覆盖率
        sheet.cell(row = i+2, column = 14).value = pubdb.DC[i]  #判定覆盖率
        sheet.cell(row = i+2, column = 15).value = pubdb.SCC[i]  #简单条件覆盖率
        sheet.cell(row = i+2, column = 16).value = pubdb.MCDC[i]  #修改条件/判定覆盖率

    #
    #NA_sheet = wb.create_sheet("NA 函数 ")

    wb.save('LRTSW-CI-子系统模块详细设计与单元测试函数的追溯表.xlsx')
    pass


def  countTestcasenum(dump):
    pubdb = Savedata()
    if True == os.path.exists('report.txt'):
        with open('report.txt',encoding='gbk') as f:
            lines = f.readlines() 
            tfilename = ""
            for line in lines:
                #匹配.c文件
                c_count = 0
                cfilename = cfilepattern.search(line)
                if cfilename is not None:
                    tfilename = cfilename.group()
                else:
                    pubdb.cname.append(tfilename)  #.c  name
                    functionname =  functionpattern.search(line)  #function name
                    pubdb.functionname.append(functionname.group())  #.c  name
                    functionname =  notDumppattern.search(functionname.group()) #notDumppattern
                    functionname = functionname.group(0)
        
                    cppfilename =  "test_"+functionname+".cpp"
                    pubdb.cppfilename.append(cppfilename)
                    #打开测试文件计算 测试案例个数
                    if True == os.path.exists(cppfilename):
                        case_num = count_test_case_num(functionname, cppfilename )
                        print("%s  num:%d"%(functionname,case_num ))
                        pubdb.testcasenum.append(case_num)
                    else:
                        print("test_"+functionname+".cpp"+"  is not exist===>>check")
                        pubdb.testcasenum.append(0)  #test case num
                    
                    #cppfilename
                    notcfileline = linepattern.search(line)
                    percentline = notcfileline.group()
                    percentline  = percentline.split('(')
                    #print(percentline)
                    parser_percent_data( percentline[0],pubdb)
    else:
        print("countTestcase-->file  report.txt  not exist")
    pubdb.mprint()
    create_excel(pubdb)